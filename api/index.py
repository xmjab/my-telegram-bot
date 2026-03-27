import os
import re
import requests
import openpyxl
from flask import Flask, request
from fpdf import FPDF

app = Flask(__name__)
TOKEN = os.environ.get('TOKEN')

# --- FUNGSI HELPER TELEGRAM ---
def send_message(chat_id, text):
    requests.post(f"https://api.telegram.org/bot{TOKEN}/sendMessage", 
                  json={"chat_id": chat_id, "text": text})

def send_document(chat_id, file_path, caption):
    url = f"https://api.telegram.org/bot{TOKEN}/sendDocument"
    with open(file_path, 'rb') as f:
        requests.post(url, data={"chat_id": chat_id, "caption": caption}, files={"document": f})

# --- LOGIKA FITUR 1: GAMAS ---
def proses_gamas(chat_id, text):
    # Logika ekstraksi STO & ODC sesuai code asli
    odp_raw = text
    odc = re.sub(r'\d', '', odp_raw.replace("ODP", "ODC").replace("/", ""))
    try: sto = odc.split("-")[1][:3]
    except: sto = "..."
    
    output = (
        f"#request\nSTO : {sto}\nDatek Terdampak (ODC): {odc}\n"
        f"Datek Terdampak (ODP): {odp_raw}\nKeterangan: ODC LOSS\n"
        f"Segmentasi: ODC\nPenyebab: Sedang Dicek\nEstimasi: 1 hari\n"
        f"PIC: yusuf 08112229796\nClassification: Z_NEW_MANUAL_GAMAS_002_004_001_004"
    )
    send_message(chat_id, output)

# --- LOGIKA FITUR 2: BA MANUAL (OUTPUT PDF) ---
def proses_ba_pdf(chat_id, raw_text):
    try:
        # 1. Parsing Data
        data = {}
        for line in raw_text.split('\n'):
            if ':' in line:
                key, val = line.split(':', 1)
                data[key.strip().upper()] = val.strip()

        # 2. Logika WH -> ID Gudang (VLOOKUP dari Excel)
        template_path = os.path.join(os.getcwd(), 'assets', 'template.xlsx')
        id_gudang_val = data.get("WH", "")
        if os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
            if "code gudang" in wb.sheetnames:
                ws_code = wb["code gudang"]
                for row in ws_code.iter_rows(min_row=1, max_col=2):
                    if row[1].value and row[1].value.strip().upper() == id_gudang_val.upper():
                        id_gudang_val = f"{row[0].value} {row[1].value}"
                        break

        # 3. Buat PDF (FPDF) - Menjaga tampilan tetap rapi
        pdf = FPDF()
        pdf.add_page()
        
        # Jika ada logo, letakkan di assets/logo.png
        # pdf.image('assets/logo.png', 10, 8, 33) 
        
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(200, 10, txt="BERITA ACARA PENGAMBILAN MATERIAL", ln=True, align='C')
        pdf.ln(10)
        
        pdf.set_font("Arial", size=10)
        pdf.cell(100, 8, txt=f"ID GUDANG : {id_gudang_val}", ln=True)
        pdf.cell(100, 8, txt=f"TANGGAL   : {data.get('TGL', '-')}", ln=True)
        pdf.cell(100, 8, txt=f"LOKASI    : {data.get('LOKASI', '-')}", ln=True)
        pdf.cell(100, 8, txt=f"MITRA     : {data.get('MITRA', '-')}", ln=True)
        pdf.ln(5)

        # Tabel Material
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(10, 8, "No", 1)
        pdf.cell(120, 8, "Nama Material", 1)
        pdf.cell(20, 8, "Satuan", 1)
        pdf.cell(20, 8, "Qty", 1)
        pdf.ln()

        pdf.set_font("Arial", size=9)
        materials = re.findall(r'-\s*(.*?)\s*=\s*(\d+)', raw_text)
        for i, (nama, qty) in enumerate(materials[:12]):
            # Logika Satuan sesuai code asli
            nama_upper = nama.upper()
            if "AC-OF-SM-ADSS" in nama_upper: satuan = "Meter"
            elif any(x in nama_upper for x in ["PU-S7.0-400NM", "PU-S9.0-140"]): satuan = "Batang"
            else: satuan = "Pcs"
            
            pdf.cell(10, 8, str(i+1), 1)
            pdf.cell(120, 8, nama[:60], 1)
            pdf.cell(20, 8, satuan, 1)
            pdf.cell(20, 8, qty, 1)
            pdf.ln()

        # Simpan & Kirim
        project_name = data.get("LOKASI", "BA_Manual").replace(" ", "_")
        pdf_file = f"/tmp/BA_{project_name}.pdf"
        pdf.output(pdf_file)
        
        send_document(chat_id, pdf_file, f"✅ PDF Berhasil Dibuat untuk {project_name}")
        os.remove(pdf_file)

    except Exception as e:
        send_message(chat_id, f"❌ Error: {str(e)}")

# --- ROUTING WEBHOOK ---
@app.route('/webhook', methods=['POST'])
def webhook():
    data = request.get_json()
    if "message" in data and "text" in data["message"]:
        chat_id = data["message"]["chat"]["id"]
        text = data["message"]["text"]

        if text == "/start":
            send_message(chat_id, "Hailurrdeeee! Bot Aktif.\n1. Kirim ODP (Contoh: ODP-STO-ABC) untuk Gamas.\n2. Kirim Format WH: untuk BA Manual.")
        
        elif text.upper().startswith("ODP"):
            proses_gamas(chat_id, text)
            
        elif "WH:" in text.upper():
            send_message(chat_id, "⏳ Memproses PDF...")
            proses_ba_pdf(chat_id, text)
            
    return "ok", 200

@app.route('/')
def index():
    return "Bot Server Running", 200